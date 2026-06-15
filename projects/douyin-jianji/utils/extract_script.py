#!/usr/bin/env python3
"""
抖音脚本Excel提取工具
从Excel文件中提取视频脚本内容并生成标准格式

使用方法:
    python3 extract_script.py <excel_file> [--output <output_dir>]
    python3 extract_script.py --source walle [--output <output_dir>]
"""

import os
import sys
import argparse
import subprocess
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要安装openpyxl: pip install openpyxl")
    sys.exit(1)


def copy_from_walle(output_dir: str = ".") -> str:
    """从walle电脑复制Excel文件"""
    source = "zhangyanbo@192.168.28.92:/home/zhangyanbo/桌面/douyinwenanjiaoben.xlsx"
    target = os.path.join(output_dir, "douyinwenanjiaoben.xlsx")

    print(f"正在从walle复制文件...")
    cmd = ["sshpass", "-p", "zyb123456", "scp", source, target]
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        raise Exception(f"复制失败: {result.stderr}")

    print(f"文件已复制到: {target}")
    return target


def extract_excel_content(excel_path: str) -> dict:
    """提取Excel内容"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    content = {
        "title": "",
        "duration": "",
        "style": "",
        "bgm": "",
        "headline": "",
        "hashtags": "",
        "camera": "",
        "sections": [],
        "execution_tips": {},
        "raw_data": []
    }

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_data = [cell.value for cell in row]
        content["raw_data"].append(row_data)

        # 提取基本信息
        if row_data[0] and "时长" in str(row_data[0]):
            content["duration"] = row_data[1] if row_data[1] else ""
        elif row_data[0] and "风格" in str(row_data[0]):
            content["style"] = row_data[1] if row_data[1] else ""
        elif row_data[0] and "BGM" in str(row_data[0]):
            content["bgm"] = row_data[1] if row_data[1] else ""
        elif row_data[0] and "标题" in str(row_data[0]):
            content["headline"] = row_data[1] if row_data[1] else ""
        elif row_data[0] and "话题" in str(row_data[0]):
            content["hashtags"] = row_data[1] if row_data[1] else ""
        elif row_data[0] and "出镜" in str(row_data[0]):
            content["camera"] = row_data[1] if row_data[1] else ""

        # 提取脚本段落
        if row_data[0] and "-" in str(row_data[0]) and "s" in str(row_data[0]):
            section = {
                "time": str(row_data[0]).strip(),
                "visual": str(row_data[1]).strip() if row_data[1] else "",
                "narration": str(row_data[2]).strip() if row_data[2] else "",
                "素材建议": str(row_data[4]).strip() if len(row_data) > 4 and row_data[4] else ""
            }
            content["sections"].append(section)

        # 提取执行建议
        if row_data[0] and "评论区预埋" in str(row_data[0]):
            content["execution_tips"]["comments"] = str(row_data[1]).strip() if row_data[1] else ""
        elif row_data[0] and "字幕细节" in str(row_data[0]):
            content["execution_tips"]["subtitle_tips"] = str(row_data[1]).strip() if row_data[1] else ""
        elif row_data[0] and "封面文案" in str(row_data[0]):
            content["execution_tips"]["cover_text"] = str(row_data[1]).strip() if row_data[1] else ""

    # 提取标题（第一行）
    if content["raw_data"] and content["raw_data"][0][0]:
        content["title"] = str(content["raw_data"][0][0]).strip()

    return content


def generate_script_md(content: dict) -> str:
    """生成Markdown格式的脚本"""
    lines = []

    # 标题
    lines.append(f"# {content['title']}")
    lines.append("")

    # 基本信息
    lines.append("## 基本信息")
    lines.append("")
    lines.append(f"- **时长**: {content['duration']}")
    lines.append(f"- **风格**: {content['style']}")
    lines.append(f"- **BGM**: {content['bgm']}")
    lines.append(f"- **标题**: {content['headline']}")
    lines.append(f"- **话题**: {content['hashtags']}")
    lines.append(f"- **出镜**: {content['camera']}")
    lines.append("")

    # 脚本内容
    lines.append("## 脚本内容")
    lines.append("")

    for i, section in enumerate(content["sections"], 1):
        lines.append(f"### {section['time']}")
        lines.append("")
        lines.append(f"**画面内容:**")
        lines.append(f"> {section['visual']}")
        lines.append("")
        lines.append(f"**配音/解说:**")
        lines.append(f"> {section['narration']}")
        lines.append("")
        if section['素材建议']:
            lines.append(f"**素材建议:** {section['素材建议']}")
            lines.append("")

    # 执行建议
    if content["execution_tips"]:
        lines.append("## 执行建议")
        lines.append("")

        if "comments" in content["execution_tips"]:
            lines.append("### 评论区预埋")
            lines.append("")
            lines.append(content["execution_tips"]["comments"])
            lines.append("")

        if "subtitle_tips" in content["execution_tips"]:
            lines.append("### 字幕细节")
            lines.append("")
            lines.append(content["execution_tips"]["subtitle_tips"])
            lines.append("")

        if "cover_text" in content["execution_tips"]:
            lines.append("### 封面文案")
            lines.append("")
            lines.append(content["execution_tips"]["cover_text"])
            lines.append("")

    return "\n".join(lines)


def generate_narration_txt(content: dict) -> str:
    """生成纯配音文本"""
    lines = []
    lines.append(f"【{content['title']}】")
    lines.append("")

    for section in content["sections"]:
        lines.append(f"【{section['time']}】")
        lines.append(section['narration'])
        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="抖音脚本Excel提取工具")
    parser.add_argument("excel_file", nargs="?", help="Excel文件路径")
    parser.add_argument("--source", choices=["walle"], help="从指定设备获取文件")
    parser.add_argument("--output", "-o", default="./scripts", help="输出目录")

    args = parser.parse_args()

    # 确定Excel文件路径
    excel_path = args.excel_file

    if args.source == "walle":
        os.makedirs(args.output, exist_ok=True)
        excel_path = copy_from_walle(args.output)
    elif not excel_path:
        parser.error("请指定Excel文件路径或使用 --source walle")

    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    print(f"正在提取: {excel_path}")
    content = extract_excel_content(excel_path)

    # 生成输出文件名
    base_name = Path(excel_path).stem

    # 确保输出目录存在
    os.makedirs(args.output, exist_ok=True)

    # 生成Markdown脚本
    md_content = generate_script_md(content)
    md_path = os.path.join(args.output, f"{base_name}_script.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"已生成脚本: {md_path}")

    # 生成配音文本
    txt_content = generate_narration_txt(content)
    txt_path = os.path.join(args.output, f"{base_name}_narration.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt_content)
    print(f"已生成配音文本: {txt_path}")

    print("\n提取完成！")


if __name__ == "__main__":
    main()
